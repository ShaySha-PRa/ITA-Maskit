"""Excel 读写。

openpyxl 读取全部 sheet，逐 sheet 脱敏后写回，保留多 sheet 结构。
每 sheet 独立按列脱敏（_mask_dataframe）。
"""
from __future__ import annotations

from pathlib import Path

from maskit.io.csvio import _mask_dataframe
from maskit.rules.defs import RuleSet

# openpyxl 用于读/写 xlsx（保留多 sheet）
try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None


def mask_excel_file(
    input_path: str | Path,
    output_path: str | Path,
    ruleset: RuleSet,
    pepper: str | None,
    details: dict | None = None,
) -> int:
    """脱敏 Excel（全部 sheet）→ Excel，返回总处理行数。

    用 openpyxl 读取全部 sheet → 逐 sheet 转 Polars 脱敏 → 写回。
    details（可选）：传入 dict 时，记录每个 sheet 的处理信息
    （sheet 名、行数、脱敏单元格数），供 GUI 展示。
    """
    src = Path(input_path)
    dst = Path(output_path)
    if not src.exists():
        raise FileNotFoundError(f"输入文件不存在: {src}")

    if load_workbook is None:
        raise ValueError("需要安装 openpyxl 才能处理 Excel")

    try:
        wb = load_workbook(src, data_only=True)
    except Exception as exc:
        raise ValueError(f"无法读取 Excel 文件: {src} ({exc})") from exc

    if not wb.sheetnames:
        raise ValueError(f"Excel 文件无工作表: {src}")

    import polars as pl

    total = 0
    sheet_info = []
    for ws in wb.worksheets:
        # 读取 sheet 数据为 list[list]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header = [str(c) if c is not None else f"col_{i}" for i, c in enumerate(rows[0])]
        data = rows[1:]
        if not data:
            continue
        # 关键：构造 DataFrame 前把所有值转成字符串。
        # 真实 Excel 同一列常混合 datetime 值、空值、数字（如日期列 + 空行），
        # 直接构造会导致 Polars 报 "could not append value: datetime[μs] ..."
        data = [
            ["" if v is None else str(v) for v in row]
            for row in data
        ]
        df = pl.DataFrame(data, schema=header, orient="row")
        masked, masked_count = _mask_dataframe(df, ruleset, pepper)
        total += masked.height
        # 记录每个 sheet 的处理信息（供 GUI 展示）
        sheet_info.append({
            "sheet": ws.title,
            "rows": masked.height,
            "masked_cells": masked_count,
            "columns": len(header),
        })
        # 写回该 sheet
        _write_sheet(ws, header, masked)

    if details is not None:
        details["sheets"] = sheet_info
    wb.save(dst)
    return total

    wb.save(dst)
    return total


def _write_sheet(ws, header: list[str], df) -> None:
    """把脱敏后的 DataFrame 写回 openpyxl worksheet。"""
    # 先清空原内容
    ws.delete_rows(1, ws.max_row)
    # 写 header
    ws.append(header)
    # 写数据（DataFrame 全为字符串列，空值写成空串）
    for row in df.iter_rows():
        ws.append(["" if v is None else str(v) for v in row])
