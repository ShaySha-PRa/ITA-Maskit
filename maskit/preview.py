"""规则集预验证：读表格文件 → 按列统计「会脱敏哪些、命中多少」。

不产出任何文件；供 GUI「预验证」按钮与 CLI 使用。
只支持表格格式（csv/xlsx/json/jsonl/ndjson）——文本格式无列，无法按列预览。
"""
from __future__ import annotations

from pathlib import Path

from maskit.rules.defs import RuleSet


def preview_ruleset_file(
    path: str | Path,
    ruleset: RuleSet,
    pepper: str | None = None,
    person_list: set[str] | None = None,
    sample_rows: int = 500,
) -> dict:
    """预验证单个表格文件，返回结构化结果。

    返回: {path, format, sheets: [{sheet, rows, columns: [...]}],
           total_hits, total_cells}
    - columns 每项来自 engine.preview_dataframe（列/命中规则/命中数/命中率/样例）
    - total_hits / total_cells 为各 sheet 汇总
    """
    src = Path(path)
    if not src.exists():
        raise FileNotFoundError(f"输入文件不存在: {src}")
    ext = src.suffix.lower()

    if ext == ".csv":
        sheets = [{"sheet": None, **_preview_csv(src, ruleset, pepper, person_list, sample_rows)}]
    elif ext in {".xlsx", ".xls"}:
        sheets = _preview_excel(src, ruleset, pepper, person_list, sample_rows)
    elif ext in {".json", ".jsonl", ".ndjson"}:
        sheets = [{"sheet": None, **_preview_json(src, ruleset, pepper, person_list, sample_rows)}]
    else:
        raise ValueError(f"预验证仅支持表格格式（csv/xlsx/json），不支持: {ext}")

    return {
        "path": str(src),
        "format": ext.lstrip("."),
        "sheets": sheets,
        "total_hits": sum(s["total_hits"] for s in sheets),
        "total_cells": sum(s["total_cells"] for s in sheets),
    }


# --- 内部：各格式读取 + 抽样预演 ---

def _preview_df(df, ruleset, pepper, person_list, sample_rows: int) -> dict:
    """对 DataFrame 抽前 sample_rows 行做预演，返回 {rows, columns, total_hits, total_cells}。"""
    from maskit.rules.engine import preview_dataframe

    if df.height == 0:
        return {"rows": 0, "columns": [], "total_hits": 0, "total_cells": 0}
    sample = df.head(sample_rows)
    columns = preview_dataframe(sample, ruleset, pepper, person_list)
    return {
        "rows": sample.height,
        "columns": columns,
        "total_hits": sum(c["hits"] for c in columns),
        "total_cells": sum(c["total"] for c in columns),
    }


def _cast_str(df):
    """统一转字符串列，消除 datetime/数字类型差异（与 csvio 一致）。"""
    import polars as pl

    return df.with_columns([pl.col(c).cast(pl.Utf8) for c in df.columns])


def _preview_csv(src, ruleset, pepper, person_list, sample_rows: int) -> dict:
    import polars as pl

    try:
        df = pl.read_csv(src, infer_schema_length=None)
    except Exception:  # noqa: BLE001 — GBK 等编码回退
        df = pl.read_csv(src, encoding="gb18030", infer_schema_length=None)
    return _preview_df(_cast_str(df), ruleset, pepper, person_list, sample_rows)


def _preview_json(src, ruleset, pepper, person_list, sample_rows: int) -> dict:
    import polars as pl

    if src.suffix.lower() in {".jsonl", ".ndjson"}:
        df = pl.read_ndjson(src)
    else:
        try:
            df = pl.read_json(src)
        except Exception:  # noqa: BLE001 — 数组 JSON 解析失败回退 JSONL
            df = pl.read_ndjson(src)
    return _preview_df(_cast_str(df), ruleset, pepper, person_list, sample_rows)


def _preview_excel(src, ruleset, pepper, person_list, sample_rows: int) -> list[dict]:
    """Excel：逐 sheet 预演，返回 [{sheet, rows, columns, ...}, ...]。"""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ValueError("需要安装 openpyxl 才能预验证 Excel")
    import polars as pl

    wb = load_workbook(src, data_only=True)
    sheets = []
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header = [str(c) if c is not None else f"col_{i}" for i, c in enumerate(rows[0])]
        data = rows[1:]
        data = [["" if v is None else str(v) for v in row] for row in data]
        df = pl.DataFrame(data, schema=header, orient="row")
        info = _preview_df(df, ruleset, pepper, person_list, sample_rows)
        info["sheet"] = ws.title
        sheets.append(info)
    if not sheets:
        raise ValueError(f"Excel 文件无可用数据: {src}")
    return sheets
